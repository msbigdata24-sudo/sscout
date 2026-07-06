from __future__ import annotations

import asyncio
import re
from urllib.parse import urlparse

from fastapi import Body, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, Field, field_validator

from server.config import (
    BUILD_VERSION,
    MAX_EXPORT_PHONES,
    PILOT_SEED_DOMAINS,
    PORT,
    ROOT,
    SCRAPINGBEE_API_KEY,
    SCRAPINGFISH_API_KEY,
    XMLRIVER_KEY,
    XMLRIVER_USER,
    YANDEX_XML_KEY,
    YANDEX_XML_USER,
)
from server.phones import normalize_digits
from server.serp import parse_xmlriver_credentials, probe_xmlriver
from server.crawler import analyze_client_site, analyze_site_for_brief, normalize_url
from server.site_survey import SiteSurveyData, suggest_from_survey
from server.db import db
from server.pipeline import (
    _can_resume_run,
    _prepare_resume_pipeline,
    find_running_run_id,
    resume_pipeline_background,
    start_pipeline_background,
    stop_pipeline,
)

app = FastAPI(title="Сигнал-Скаут API", version="1.1")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class BriefModel(BaseModel):
    clientName: str = ""
    clientSite: str
    niche: str = ""
    regionMode: str = "include"
    regions: str = ""
    queries: str = ""
    excludeDomains: str = ""
    phoneFilter: str = "business"
    sources: list[str] = Field(default_factory=lambda: ["serp", "site", "catalog"])
    checkAlive: bool = True
    xmlRiverUser: str = ""
    apiKey: str = ""
    maxSites: int = Field(default=50, ge=10, le=200)
    crawlDepth: int = Field(default=2, ge=1, le=5)
    requestDelayMs: int = Field(default=500, ge=0, le=5000)
    useProxy: bool = False
    quickCrawl: bool = False
    seedDomains: str = ""

    @field_validator("clientSite")
    @classmethod
    def validate_client_site(cls, value: str) -> str:
        normalized = normalize_client_site(value)
        if not normalized:
            raise ValueError("Укажите корректный URL сайта, например https://example.ru")
        host = urlparse(normalized).netloc
        if "." not in host:
            raise ValueError("В адресе сайта должно быть доменное имя")
        return normalized


def normalize_client_site(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    url_match = re.search(r"https?://[^\s<>\"'·|]+", text, flags=re.IGNORECASE)
    if url_match:
        text = url_match.group(0).rstrip(".,;)")
    elif "://" not in text and re.search(
        r"[\w.-]+\.(ru|com|рф|org|net|biz)\b", text, flags=re.IGNORECASE
    ):
        text = "https://" + text.lstrip("/").rstrip(".,;)")
    return normalize_url(text)


@app.get("/api/health")
def health():
    user, key = parse_xmlriver_credentials(api_user=XMLRIVER_USER, api_key=XMLRIVER_KEY)
    return {
        "ok": True,
        "version": BUILD_VERSION,
        "branch": "main",
        "features": ["quick-crawl", "pilot-queries", "resume", "brief-suggest"],
        "search_provider": "xmlriver",
        "xmlriver_configured": bool(user and key),
        "yandex_xml_fallback": bool(YANDEX_XML_USER and YANDEX_XML_KEY),
        "scraping_configured": bool(SCRAPINGBEE_API_KEY or SCRAPINGFISH_API_KEY),
        "proxy_configured": bool(SCRAPINGBEE_API_KEY or SCRAPINGFISH_API_KEY),
        "port": PORT,
    }


@app.get("/api/xmlriver/check")
async def xmlriver_check(
    xmlRiverUser: str = Query("", alias="xmlRiverUser"),
    apiKey: str = Query("", alias="apiKey"),
):
    """Тестовый запрос к XMLRiver — показывает реальную ошибку API."""
    user = xmlRiverUser or XMLRIVER_USER
    key = apiKey or XMLRIVER_KEY
    return await probe_xmlriver(xmlriver_user=user, xmlriver_key=key)


def _search_ready(brief: BriefModel) -> bool:
    if brief.quickCrawl:
        return True
    user, key = parse_xmlriver_credentials(
        api_user=brief.xmlRiverUser or XMLRIVER_USER,
        api_key=brief.apiKey or XMLRIVER_KEY,
    )
    if user and key:
        return True
    return bool(YANDEX_XML_USER and YANDEX_XML_KEY)


@app.get("/api/preview")
async def preview_site(url: str = Query(..., min_length=4)):
    data = await analyze_client_site(url, depth=1, delay_ms=0)
    if not data.get("ok"):
        raise HTTPException(400, data.get("error") or "Не удалось открыть сайт")
    return {
        "title": data.get("title") or "",
        "niche_hint": (data.get("title") or "")[:200],
        "sample": (data.get("text_sample") or "")[:500],
    }


@app.get("/api/brief/suggest")
async def suggest_brief(url: str = Query(..., min_length=4)):
    """Опрос сайта (несколько страниц): ниша, запросы, исключения. Регионы — вручную."""
    normalized = normalize_client_site(url)
    if not normalized:
        raise HTTPException(400, "Укажите корректный URL сайта")
    data = await analyze_site_for_brief(normalized)
    if not data.get("ok"):
        raise HTTPException(400, data.get("error") or "Не удалось открыть сайт")
    survey_data = SiteSurveyData(
        site_url=data.get("site_url") or normalized,
        title=data.get("title") or "",
        meta_description=data.get("meta_description") or "",
        headings=data.get("headings") or [],
        nav_labels=data.get("nav_labels") or [],
        footer_text=data.get("footer_text") or "",
        brand_hints=data.get("brand_hints") or [],
        org_names=data.get("org_names") or [],
        body_text=data.get("text_sample") or "",
        list_items=data.get("list_items") or [],
        schema_offerings=data.get("schema_offerings") or [],
        pages_surveyed=int(data.get("pages_surveyed") or 1),
    )
    payload = suggest_from_survey(survey_data)
    return {"ok": True, **payload}


@app.post("/api/run")
async def api_start_run(brief: BriefModel):
    if not brief.clientSite.strip():
        raise HTTPException(400, "Укажите сайт клиента")
    if not _search_ready(brief):
        raise HTTPException(
            400,
            "Укажите ID и API-ключ XMLRiver в брифе (xmlriver.com) "
            "или YANDEX_XML_USER / YANDEX_XML_KEY в .env",
        )
    client_site = brief.clientSite.strip()
    mem_id = find_running_run_id(client_site)
    if mem_id:
        return {"run_id": mem_id, "status": "running", "reconnected": True, "resumed": False}

    site_key = _client_site_key(client_site)
    for it in db.list_runs(30):
        if _client_site_key(it.get("client_site") or "") != site_key:
            continue
        if it.get("status") not in ("stopped", "error", "running"):
            continue
        run = db.get_run(it["id"])
        if not run or not _can_resume_run(run):
            continue
        db.update_run(it["id"], brief=brief.model_dump())
        await resume_pipeline_background(it["id"])
        return {"run_id": it["id"], "status": "running", "resumed": True}

    try:
        run_id = await start_pipeline_background(brief.model_dump())
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    return {"run_id": run_id, "status": "pending", "resumed": False}


@app.post("/api/run/quick")
async def api_quick_run(brief: BriefModel):
    """Быстрый обход 12 известных конкурентов — без XMLRiver, чтобы проверить телефоны."""
    if not brief.clientSite.strip():
        raise HTTPException(400, "Укажите сайт клиента")
    payload = brief.model_dump()
    payload["quickCrawl"] = True
    payload["sources"] = ["site"]
    payload["checkAlive"] = False
    payload["maxSites"] = min(int(payload.get("maxSites") or 50), len(PILOT_SEED_DOMAINS))
    if find_running_run_id(payload["clientSite"]) or db.find_active_run(payload["clientSite"]):
        active = db.find_active_run(payload["clientSite"])
        run_id = active["id"] if active else find_running_run_id(payload["clientSite"])
        raise HTTPException(409, f"Сбор для этого клиента уже выполняется (прогон {run_id})")
    try:
        run_id = await start_pipeline_background(payload)
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    return {"run_id": run_id, "status": "pending", "quick": True}


PIPELINE_STEP_IDS = ("analyze", "serp", "filter", "crawl", "catalog", "dedup")


def _client_site_key(url: str) -> str:
    raw = (url or "").strip()
    if not raw:
        return ""
    try:
        host = urlparse(raw if "://" in raw else f"https://{raw}").netloc.lower()
    except Exception:
        host = raw.lower().split("/")[0]
    return host[4:] if host.startswith("www.") else host


def _calc_progress(pipeline: dict, status: str) -> tuple[int, str]:
    if status == "done":
        return 100, ""
    steps = list(PIPELINE_STEP_IDS)
    progress = 0
    current = ""
    for step in steps:
        st = pipeline.get(step, "pending")
        if st == "done":
            progress += 100 // len(steps)
        elif st == "running":
            current = step
            progress += 100 // (len(steps) * 2)
            if step == "crawl":
                site_status = pipeline.get("site_status") or {}
                total = max(1, len(site_status))
                done_sites = sum(1 for v in site_status.values() if v in ("success", "error", "skip"))
                progress += int((100 // len(steps)) * 0.5 * done_sites / total)
    return min(progress, 99 if status in ("running", "pending") else 100), current


@app.get("/api/run/{run_id}")
def api_get_run(run_id: str):
    run = db.get_run(run_id)
    if not run:
        raise HTTPException(404, "Прогон не найден")
    pipeline = run.get("pipeline") or {}
    status = run.get("status") or "pending"
    progress, current_step = _calc_progress(pipeline, status)
    return {
        "id": run["id"],
        "status": status,
        "pipeline": pipeline,
        "progress": progress,
        "current_step": current_step,
        "logs": pipeline.get("logs") or [],
        "site_status": pipeline.get("site_status") or {},
        "results_count": len(run.get("results") or []),
        "can_resume": _can_resume_run(run),
        "error": run.get("error"),
        "is_demo": run.get("is_demo", False),
        "updated_at": run.get("updated_at"),
    }


@app.post("/api/run/{run_id}/stop")
def api_stop_run(run_id: str):
    run = db.get_run(run_id)
    if not run:
        raise HTTPException(404, "Прогон не найден")
    stop_pipeline(run_id)
    pipeline = _prepare_resume_pipeline(run.get("pipeline") or {})
    db.update_run(
        run_id,
        status="stopped",
        pipeline=pipeline,
        results=run.get("results") or [],
    )
    return {"ok": True, "run_id": run_id, "can_resume": _can_resume_run({**run, "status": "stopped", "pipeline": pipeline})}


@app.post("/api/run/{run_id}/resume")
async def api_resume_run(run_id: str, brief: BriefModel | None = Body(default=None)):
    if not db.get_run(run_id):
        raise HTTPException(404, "Прогон не найден")
    try:
        if brief is not None:
            db.update_run(run_id, brief=brief.model_dump())
        await resume_pipeline_background(run_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"run_id": run_id, "status": "running", "resumed": True}


@app.get("/api/results/{run_id}")
def api_results(run_id: str):
    run = db.get_run(run_id)
    if not run:
        raise HTTPException(404, "Прогон не найден")
    return {
        "id": run_id,
        "status": run["status"],
        "brief": run["brief"],
        "results": run.get("results") or [],
        "is_demo": run.get("is_demo", False),
        "error": run.get("error"),
        "created_at": run.get("created_at"),
    }


@app.get("/api/history")
def api_history(limit: int = Query(50, ge=1, le=200)):
    return {"items": db.list_runs(limit)}


@app.get("/api/history/compare")
def api_compare(a: str = Query(...), b: str = Query(...)):
    ra, rb = db.get_run(a), db.get_run(b)
    if not ra or not rb:
        raise HTTPException(404, "Сессия не найдена")
    sa = {r["site"] for r in ra.get("results") or []}
    sb = {r["site"] for r in rb.get("results") or []}
    return {
        "a": a,
        "b": b,
        "new_sites": sorted(sb - sa),
        "removed_sites": sorted(sa - sb),
        "common": sorted(sa & sb),
    }


def _phones_for_export(row: dict) -> list[str]:
    raw = row.get("phones")
    if isinstance(raw, list) and raw:
        nums = [str(p) for p in raw if p]
    else:
        nums = [str(row.get(k)) for k in ("p1", "p2") if row.get(k)]
    cleaned = [normalize_digits(p) for p in nums if normalize_digits(p)]
    return cleaned[:MAX_EXPORT_PHONES]


def _export_table(rows: list[dict]) -> tuple[list[str], list[list[str]]]:
    max_phones = max((len(_phones_for_export(r)) for r in rows), default=0)
    max_phones = max(1, min(MAX_EXPORT_PHONES, max_phones))
    header = (
        ["Сайт", "Компания", "Регион"]
        + [f"Телефон {i}" for i in range(1, max_phones + 1)]
        + ["Источник", "Статус"]
    )
    data: list[list[str]] = []
    for r in rows:
        phones = _phones_for_export(r)
        line = [r.get("site", ""), r.get("name", ""), r.get("region", "")]
        for i in range(max_phones):
            line.append(phones[i] if i < len(phones) else "")
        line += [r.get("source", ""), r.get("status", "")]
        data.append(line)
    return header, data


def _export_rows(run_id: str) -> list[dict]:
    run = db.get_run(run_id)
    if not run:
        raise HTTPException(404, "Прогон не найден")
    return run.get("results") or []


def _safe_export_basename(run_id: str) -> str:
    run = db.get_run(run_id)
    name = ""
    if run:
        brief = run.get("brief") or {}
        name = (brief.get("clientName") or "").strip()
    if not name:
        return f"signal-scout-{run_id[:8]}"
    cleaned = re.sub(r'[<>:"/\\|?*\n\r\t]', "", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip().strip(".")
    return cleaned[:80] if cleaned else f"signal-scout-{run_id[:8]}"


def _export_filename(run_id: str, ext: str) -> str:
    from datetime import datetime, timezone, timedelta

    msk = timezone(timedelta(hours=3))
    date_str = datetime.now(msk).strftime("%d-%m-%Y")
    return f"{_safe_export_basename(run_id)} {date_str}.{ext}"


def _content_disposition(filename: str) -> str:
    ascii_name = re.sub(r"[^\w\-]+", "_", filename, flags=re.ASCII).strip("_") or "export"
    from urllib.parse import quote

    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"


@app.get("/api/export/{run_id}.csv")
def export_csv(run_id: str):
    import csv
    import io

    rows = _export_rows(run_id)
    header, data = _export_table(rows)
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf, delimiter=";")
    w.writerow(header)
    w.writerows(data)
    filename = _export_filename(run_id, "csv")
    return Response(
        content=buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


@app.get("/api/export/{run_id}.xlsx")
def export_xlsx(run_id: str):
    import io

    from openpyxl import Workbook

    rows = _export_rows(run_id)
    header, data = _export_table(rows)
    phone_col_start = 3
    phone_col_count = sum(1 for h in header if h.startswith("Телефон"))
    text_col_indices = set(range(phone_col_start, phone_col_start + phone_col_count))
    text_col_indices.add(len(header) - 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сигнал-Скаут"
    ws.append(header)
    for row in data:
        ws.append(row)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in text_col_indices:
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.number_format = "@"

    buf = io.BytesIO()
    wb.save(buf)
    filename = _export_filename(run_id, "xlsx")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">'
        '<rect width="32" height="32" rx="7" fill="#3d8bfd"/>'
        '<path d="M16 7l9 16H7z" fill="#fff" opacity=".95"/>'
        "</svg>"
    )
    return Response(content=svg, media_type="image/svg+xml")


@app.get("/")
def index_page():
    index = ROOT / "index.html"
    if not index.exists():
        raise HTTPException(404, "index.html не найден")
    html = index.read_text(encoding="utf-8")
    v = BUILD_VERSION
    html = html.replace("/static/js/storage.js", f"/static/js/storage.js?v={v}")
    html = html.replace("/static/js/regions-ru.js", f"/static/js/regions-ru.js?v={v}")
    html = html.replace("/static/js/app.js", f"/static/js/app.js?v={v}")
    return Response(
        content=html,
        media_type="text/html; charset=utf-8",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/static/{path:path}")
def static_file(path: str):
    f = ROOT / "static" / path
    if not f.exists() or not f.is_file():
        raise HTTPException(404)
    headers = {}
    if path.endswith(".js") or path.endswith(".css"):
        headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return FileResponse(f, headers=headers)
